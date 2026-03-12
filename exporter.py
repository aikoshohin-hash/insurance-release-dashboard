"""CSV/Excel 出力モジュール v3 - スコアリング・分析対応"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, CATEGORY_LABELS

logger = logging.getLogger(__name__)


def _ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── v3 拡張カラム ──
COLUMNS_V3 = [
    "ランク", "スコア", "人気度", "会社名", "日付",
    "見出し", "URL", "商品タイプ", "コメンタリー", "抽出理由",
]

# 旧カラム（後方互換）
COLUMNS = ["会社名", "日付", "見出し", "URL", "抽出理由"]


def _stars_text(n: int) -> str:
    """★を文字列で表現"""
    return "★" * n + "☆" * (5 - n)


def export_categorized_excel(
    categorized: dict[str, list[dict]],
    filename: str | None = None,
    enhanced: bool = True,
) -> str:
    """カテゴリ別 Excel 出力

    Args:
        categorized: {"A": [...], "B": [...], "C": [...]}
        filename: 出力ファイル名
        enhanced: True の場合、スコア・分析列も含める（v3形式）
    """
    _ensure_output_dir()
    if filename is None:
        filename = f"releases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    from openpyxl import Workbook
    wb = Workbook()

    # ── シート1: カテゴリ別一覧 ──
    ws = wb.active
    ws.title = "カテゴリ別一覧"

    columns = COLUMNS_V3 if enhanced else COLUMNS
    num_cols = len(columns)

    # スタイル定義
    cat_font = Font(bold=True, size=14, color="FFFFFF")
    cat_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ランク用スタイル
    rank_fills = {
        "S": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "A": PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
        "B": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
        "C": PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid"),
        "D": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
    }
    rank_fonts = {
        "S": Font(bold=True, size=12, color="CC0000"),
        "A": Font(bold=True, size=11, color="CC6600"),
        "B": Font(bold=True, size=11, color="228822"),
        "C": Font(size=11, color="3366CC"),
        "D": Font(size=10, color="888888"),
    }

    row_idx = 1
    total_count = 0

    for cat_key in ("A", "B", "C"):
        cat_label = CATEGORY_LABELS.get(cat_key, cat_key)
        entries = categorized.get(cat_key, [])

        # カテゴリタイトル行
        ws.merge_cells(
            start_row=row_idx, start_column=1,
            end_row=row_idx, end_column=num_cols,
        )
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"【カテゴリ{cat_key}】{cat_label}（{len(entries)}件）"
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal="left")
        row_idx += 1

        # カラムヘッダー行
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
        row_idx += 1

        # データ行
        if entries:
            for e in entries:
                rank = e.get("rank_label", "D")

                if enhanced:
                    vals = [
                        rank,
                        e.get("score", 0),
                        _stars_text(e.get("popularity", 1)),
                        e.get("company", ""),
                        e.get("date", ""),
                        e.get("title", ""),
                        e.get("url", ""),
                        e.get("product_type", ""),
                        e.get("commentary", ""),
                        e.get("reason", ""),
                    ]
                else:
                    vals = [
                        e.get("company", ""),
                        e.get("date", ""),
                        e.get("title", ""),
                        e.get("url", ""),
                        e.get("reason", ""),
                    ]

                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

                # ランク行の背景色
                if enhanced and rank in rank_fills:
                    for col_idx in range(1, num_cols + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = rank_fills[rank]
                    # ランクセルのフォント
                    ws.cell(row=row_idx, column=1).font = rank_fonts.get(rank, Font())

                row_idx += 1
                total_count += 1
        else:
            ws.cell(row=row_idx, column=1, value="（該当なし）")
            row_idx += 1

        row_idx += 1  # 空行

    # 列幅調整
    if enhanced:
        col_widths = [8, 8, 14, 18, 14, 55, 60, 18, 50, 30]
    else:
        col_widths = [18, 14, 60, 70, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── シート2: 総合ランキング ──
    if enhanced:
        ws2 = wb.create_sheet("総合ランキング")
        _write_ranking_sheet(ws2, categorized, rank_fills, rank_fonts, thin_border)

    # ── シート3: サマリー ──
    if enhanced:
        ws3 = wb.create_sheet("サマリー")
        _write_summary_sheet(ws3, categorized, cat_fill, cat_font, header_fill, header_font)

    wb.save(filepath)
    logger.info(f"Excel出力完了: {filepath} ({total_count}件)")
    return filepath


def _write_ranking_sheet(ws, categorized, rank_fills, rank_fonts, thin_border):
    """総合ランキングシートを作成"""
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5496")

    # タイトル
    ws.merge_cells("A1:J1")
    cell = ws.cell(row=1, column=1, value="📊 保険リリース 総合ランキング")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"作成日: {datetime.now().strftime('%Y/%m/%d %H:%M')}")

    # 全エントリ統合・スコア順
    all_entries = []
    for cat, entries in categorized.items():
        for e in entries:
            e_copy = e.copy()
            e_copy["cat_key"] = cat
            all_entries.append(e_copy)
    all_entries.sort(key=lambda x: x.get("score", 0), reverse=True)

    # ヘッダー
    rank_headers = [
        "順位", "ランク", "スコア", "人気度", "会社名",
        "日付", "見出し", "カテゴリ", "商品タイプ", "コメンタリー",
    ]
    row_idx = 4
    for col_idx, h in enumerate(rank_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row_idx += 1

    # データ
    for i, e in enumerate(all_entries, 1):
        rank = e.get("rank_label", "D")
        vals = [
            i,
            rank,
            e.get("score", 0),
            _stars_text(e.get("popularity", 1)),
            e.get("company", ""),
            e.get("date", ""),
            e.get("title", ""),
            f"{e.get('cat_key', '')}:{CATEGORY_LABELS.get(e.get('cat_key', ''), '')}",
            e.get("product_type", ""),
            e.get("commentary", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if rank in rank_fills:
                cell.fill = rank_fills[rank]
        if rank in rank_fonts:
            ws.cell(row=row_idx, column=2).font = rank_fonts[rank]
        row_idx += 1

    # 列幅
    widths = [8, 8, 8, 14, 18, 14, 55, 18, 18, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(ws, categorized, cat_fill, cat_font, header_fill, header_font):
    """サマリーシートを作成"""
    title_font = Font(bold=True, size=16, color="2F5496")
    sub_font = Font(bold=True, size=12, color="2F5496")

    # タイトル
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="📋 サマリーレポート")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"作成日: {datetime.now().strftime('%Y/%m/%d %H:%M')}")

    row_idx = 4

    # カテゴリ別サマリー
    ws.cell(row=row_idx, column=1, value="■ カテゴリ別集計").font = sub_font
    row_idx += 1
    for cat in ("A", "B", "C"):
        entries = categorized.get(cat, [])
        ws.cell(row=row_idx, column=1, value=f"カテゴリ{cat} ({CATEGORY_LABELS.get(cat, '')})")
        ws.cell(row=row_idx, column=2, value=f"{len(entries)}件")
        if entries:
            scores = [e.get("score", 0) for e in entries]
            ws.cell(row=row_idx, column=3, value=f"平均スコア: {sum(scores)/len(scores):.1f}")
        row_idx += 1

    row_idx += 1

    # ランク別集計
    ws.cell(row=row_idx, column=1, value="■ ランク別集計").font = sub_font
    row_idx += 1
    all_entries = [e for entries in categorized.values() for e in entries]
    for rank in ("S", "A", "B", "C", "D"):
        count = sum(1 for e in all_entries if e.get("rank_label") == rank)
        ws.cell(row=row_idx, column=1, value=f"ランク {rank}")
        ws.cell(row=row_idx, column=2, value=f"{count}件")
        row_idx += 1

    row_idx += 1

    # 会社別集計
    ws.cell(row=row_idx, column=1, value="■ 会社別集計").font = sub_font
    row_idx += 1
    company_data = {}
    for e in all_entries:
        co = e.get("company", "不明")
        if co not in company_data:
            company_data[co] = {"count": 0, "scores": []}
        company_data[co]["count"] += 1
        company_data[co]["scores"].append(e.get("score", 0))

    for co, info in sorted(company_data.items(), key=lambda x: -x[1]["count"]):
        avg = sum(info["scores"]) / len(info["scores"]) if info["scores"] else 0
        ws.cell(row=row_idx, column=1, value=co)
        ws.cell(row=row_idx, column=2, value=f"{info['count']}件")
        ws.cell(row=row_idx, column=3, value=f"平均スコア: {avg:.1f}")
        row_idx += 1

    row_idx += 1

    # 商品タイプ集計
    ws.cell(row=row_idx, column=1, value="■ 商品タイプ別集計").font = sub_font
    row_idx += 1
    ptype_counts = {}
    for e in all_entries:
        pt = e.get("product_type", "")
        if pt:
            ptype_counts[pt] = ptype_counts.get(pt, 0) + 1
    for pt, cnt in sorted(ptype_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row_idx, column=1, value=pt)
        ws.cell(row=row_idx, column=2, value=f"{cnt}件")
        row_idx += 1

    # 列幅
    for i, w in enumerate([25, 12, 20, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_csv(releases: list[dict], filename: str | None = None) -> str:
    """全データ CSV 出力（後方互換）"""
    _ensure_output_dir()
    if filename is None:
        filename = f"releases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(OUTPUT_DIR, filename)

    df = pd.DataFrame(releases)
    if not df.empty:
        df = df.sort_values(["company", "date"], ascending=[True, False])
    df.to_csv(filepath, index=False, encoding="utf-8-sig")

    logger.info(f"CSV出力完了: {filepath} ({len(df)}件)")
    return filepath
